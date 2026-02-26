"""
Модуль для автоматической генерации выходного Excel файла.
"""
from typing import Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

from src.models import AppConfig, CompetitorConfig

logger = logging.getLogger(__name__)


class OutputFileGenerator:
    """Генератор выходного Excel файла."""

    # Поля конкурента, которые нужно вывести
    FIELDS = ['convert', 'minimum_1', 'minimum_2', 'volume', 'weight_100', 'weight_3000']
    FIELD_NAMES = {
        'convert': 'Конверт',
        'minimum_1': 'Посылка до 10 кг',
        'minimum_2': '1 место до 30 кг',
        'volume': 'Груз до 0,5 куба',
        'weight_100': 'Груз до 100 кг',
        'weight_3000': 'Груз более 3000 кг'
    }

    def __init__(self, config: AppConfig):
        self.config = config
        self.wb = None
        self.ws = None
        self.data: Dict[str, Dict[str, Dict[str, Any]]] = {}
        self.row_map: Dict[str, Dict[str, int]] = {}
        # строка на листе «Наценки» для каждого конкурента: {имя: строка}
        self.markups_row_map: Dict[str, int] = {}

    def generate(self, city_competitors: Dict[str, list] = None) -> bool:
        """Полностью сгенерировать выходной файл."""
        try:
            # Создать новую книгу
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.title = "Данные"

            # Инициализировать структуру данных
            self.data = {city: {} for city in self.config.cities.keys()}

            # Создать заголовки
            self._create_headers()

            # Создать пустые строки для данных
            self._create_empty_rows(city_competitors=city_competitors)

            # Сохранить файл
            if not self.config.output_file:
                logger.error("Не указан путь для выходного файла")
                return False

            self.wb.save(self.config.output_file)
            logger.info(f"Выходной файл создан: {self.config.output_file}")

            return True

        except Exception as e:
            logger.error(f"Ошибка генерации файла: {e}")
            return False

    def _create_headers(self):
        """Создать общий заголовок файла (строка 1)."""
        out_cfg = self.config.output_config

        # Строка 1 — общий заголовок, объединённая
        title_cell = self.ws['A1']
        title_cell.value = out_cfg.title
        if out_cfg.subtitle:
            title_cell.value += f" — {out_cfg.subtitle}"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        total_cols = 1 + len(self.FIELDS)
        self.ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")

        # Ширина колонок (задаётся один раз)
        self.ws.column_dimensions[get_column_letter(1)].width = 22
        for col_idx in range(2, 2 + len(self.FIELDS)):
            self.ws.column_dimensions[get_column_letter(col_idx)].width = 18

    def _write_column_headers(self, row: int):
        """Нарисовать строку заголовков колонок на указанной строке."""
        out_cfg = self.config.output_config

        # Первая колонка — название таблицы
        cell = self.ws.cell(row=row, column=1)
        cell.value = out_cfg.title
        self._style_header_cell(cell)

        # Колонки полей
        for fi, field in enumerate(self.FIELDS):
            cell = self.ws.cell(row=row, column=2 + fi)
            cell.value = self.FIELD_NAMES[field]
            self._style_header_cell(cell)


    def _create_empty_rows(self, city_competitors: Dict[str, list] = None):
        """Создать строки для городов, конкурентов, среднего значения и собственной ТК."""
        enabled_competitors = [
            c for c in self.config.competitors.values()
            if c.enabled
        ]
        own = self.config.own_company

        self.row_map = {}
        # Данные начинаются со строки 2 (строка 1 — общий заголовок)
        current_row = 2

        # Города строго в алфавитном порядке
        sorted_cities = sorted(self.config.cities.keys())

        for city_idx, city in enumerate(sorted_cities):
            self.row_map[city] = {}

            # Конкуренты для этого города: если передана карта — берём из неё,
            # иначе все включённые (обратная совместимость)
            if city_competitors is not None:
                competitors_for_city = city_competitors.get(city, [])
            else:
                competitors_for_city = enabled_competitors

            # --- Строка заголовков колонок для каждого города ---
            self._write_column_headers(current_row)
            current_row += 1

            # --- Строка города (жирная, цветная) ---
            city_cell = self.ws.cell(row=current_row, column=1)
            city_cell.value = city
            self._style_city_cell(city_cell)
            for col_idx in range(2, 2 + len(self.FIELDS)):
                self._style_city_cell(self.ws.cell(row=current_row, column=col_idx))
            current_row += 1

            # --- Строки конкурентов (только те, у которых есть данные для города) ---
            first_competitor_row = current_row
            for competitor in competitors_for_city:
                self.row_map[city][competitor.name] = current_row
                name_cell = self.ws.cell(row=current_row, column=1)
                name_cell.value = competitor.name
                self._style_data_cell(name_cell, bold=competitor.bold)
                for col_idx in range(2, 2 + len(self.FIELDS)):
                    self._style_data_cell(self.ws.cell(row=current_row, column=col_idx), bold=competitor.bold)
                current_row += 1

                # Строки наценок этого конкурента
                for mk_row in competitor.markup_rows:
                    mk_key = f"{competitor.name}|{mk_row.name}"
                    self.row_map[city][mk_key] = current_row
                    mk_name_cell = self.ws.cell(row=current_row, column=1)
                    mk_name_cell.value = mk_row.name
                    self._style_markup_row_cell(mk_name_cell)
                    for col_idx in range(2, 2 + len(self.FIELDS)):
                        self._style_markup_row_cell(self.ws.cell(row=current_row, column=col_idx))
                    current_row += 1

            last_competitor_row = current_row - 1

            # --- Строка «Среднее значение» ---
            if self.config.output_config.include_average:
                avg_row = current_row
                self.row_map[city]["__average__"] = avg_row
                avg_name_cell = self.ws.cell(row=avg_row, column=1)
                avg_name_cell.value = "Среднее значение"
                self._style_average_cell(avg_name_cell)

                for fi in range(len(self.FIELDS)):
                    col_idx = 2 + fi
                    col_letter = get_column_letter(col_idx)
                    cell = self.ws.cell(row=avg_row, column=col_idx)
                    if competitors_for_city:
                        cell.value = (
                            f"=AVERAGE({col_letter}{first_competitor_row}"
                            f":{col_letter}{last_competitor_row})"
                        )
                    self._style_average_cell(cell)

                current_row += 1

                # --- Строка собственной ТК ---
                if own.enabled:
                    self.row_map[city]["__own__"] = current_row
                    own_name_cell = self.ws.cell(row=current_row, column=1)
                    own_name_cell.value = own.name
                    self._style_own_cell(own_name_cell)
                    for col_idx in range(2, 2 + len(self.FIELDS)):
                        self._style_own_cell(self.ws.cell(row=current_row, column=col_idx))
                    current_row += 1

            # --- Пустая строка-разделитель между городами ---
            if city_idx < len(sorted_cities) - 1:
                current_row += 1

    def _style_header_cell(self, cell):
        """Стилизовать ячейку заголовка."""
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="385E72", end_color="385E72", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _style_city_cell(self, cell):
        """Стилизовать ячейку названия города."""
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="D9E4EC", end_color="D9E4EC", fill_type="solid")
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _style_data_cell(self, cell, bold: bool = False):
        """Стилизовать ячейку данных."""
        cell.font = Font(bold=bold, size=11)
        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _style_average_cell(self, cell):
        """Стилизовать ячейку среднего значения."""
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _style_own_cell(self, cell):
        """Стилизовать ячейку собственной ТК (Новая Витэка)."""
        cell.font = Font(bold=False, size=11, color="1F497D")
        cell.fill = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _style_markup_row_cell(self, cell):
        """Стилизовать ячейку строки наценки конкурента."""
        cell.fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def write_competitor_data(
        self,
        competitor: CompetitorConfig,
        city: str,
        field: str,
        value: Any
    ):
        """Записать данные конкурента в файл."""
        if city not in self.row_map:
            return
        if not self.wb:
            return
        if competitor.name not in self.row_map[city]:
            return
        if field not in self.FIELDS:
            return

        col_idx = 2 + self.FIELDS.index(field)
        col_letter = get_column_letter(col_idx)
        row = self.row_map[city][competitor.name]

        cell = self.ws.cell(row=row, column=col_idx)

        if value is not None and isinstance(value, (int, float)):
            markup_percent = getattr(competitor.markups, field, 0)
            if markup_percent != 0 and competitor.name in self.markups_row_map:
                # Ссылаемся на ячейку процента с листа «Наценки»
                markup_row = self.markups_row_map[competitor.name]
                markup_cell_ref = f"Наценки!{col_letter}{markup_row}"
                cell.value = f"={value}*(1+{markup_cell_ref}/100)"
            else:
                cell.value = value
        else:
            cell.value = value

        # Строки наценок (markup_rows) — ссылаются на ячейку конкурента
        for mk_row in competitor.markup_rows:
            mk_key = f"{competitor.name}|{mk_row.name}"
            if mk_key not in self.row_map[city]:
                continue
            mk_row_num = self.row_map[city][mk_key]
            mk_cell = self.ws.cell(row=mk_row_num, column=col_idx)
            if mk_row.percent != 0:
                mk_cell.value = f"={col_letter}{row}*(1+{mk_row.percent}/100)"
            else:
                mk_cell.value = f"={col_letter}{row}"

        # Обновить в памяти
        if competitor.name not in self.data[city]:
            self.data[city][competitor.name] = {}
        self.data[city][competitor.name][field] = value

    def _find_column(self, competitor: CompetitorConfig, field: str) -> Optional[int]:
        """Найти номер колонки для поля в новой структуре."""
        if field not in self.FIELDS:
            return None
        return 2 + self.FIELDS.index(field)

    def save(self) -> bool:
        """Сохранить файл."""
        try:
            if not self.wb:
                return False

            self.wb.save(self.config.output_file)
            logger.info(f"Файл сохранен: {self.config.output_file}")
            return True
        except Exception as e:
            logger.error(f"Ошибка сохранения файла: {e}")
            return False

    def add_markups_sheet(self):
        """Добавить лист «Наценки» с редактируемыми процентами."""
        if not self.config.output_config.markups_sheet:
            return

        try:
            markups_ws = self.wb.create_sheet("Наценки")

            # Заголовок
            title_cell = markups_ws['A1']
            title_cell.value = "Наценки на цены конкурентов (%)"
            title_cell.font = Font(size=14, bold=True)
            total_cols = 1 + len(self.FIELDS)
            markups_ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")

            # Строка заголовков колонок (строка 3)
            headers = ["Конкурент"] + [self.FIELD_NAMES[f] for f in self.FIELDS]
            for col_idx, header in enumerate(headers, 1):
                cell = markups_ws.cell(row=3, column=col_idx)
                cell.value = header
                self._style_header_cell(cell)

            # Данные наценок начиная со строки 4
            enabled_competitors = [
                c for c in self.config.competitors.values()
                if c.enabled
            ]

            self.markups_row_map = {}
            for row_idx, competitor in enumerate(enabled_competitors, 4):
                self.markups_row_map[competitor.name] = row_idx

                # Название конкурента
                cell = markups_ws.cell(row=row_idx, column=1)
                cell.value = competitor.name
                self._style_city_cell(cell)

                # Проценты наценок — числа, чтобы на них можно было ссылаться
                for fi, field in enumerate(self.FIELDS):
                    col_idx = 2 + fi
                    cell = markups_ws.cell(row=row_idx, column=col_idx)
                    markup_value = getattr(competitor.markups, field, 0)
                    cell.value = markup_value  # число, не строка!
                    # Формат отображения со знаком %
                    cell.number_format = '0.##"%"'
                    self._style_data_cell(cell)

            # Ширина колонок
            markups_ws.column_dimensions['A'].width = 22
            for col_idx in range(2, 2 + len(self.FIELDS)):
                markups_ws.column_dimensions[get_column_letter(col_idx)].width = 15

            logger.info("Лист с наценками добавлен")

        except Exception as e:
            logger.error(f"Ошибка добавления листа наценок: {e}")
