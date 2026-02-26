"""
Модуль для работы с Excel файлами.
"""
from pathlib import Path
from typing import Dict, List, Optional, Any
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from thefuzz import fuzz
import logging
from openpyxl.utils import get_column_letter

from src.models import CompetitorConfig, AppConfig
from src.output_generator import OutputFileGenerator

logger = logging.getLogger(__name__)


class ExcelProcessor:
    """Класс для обработки Excel файлов."""

    def __init__(self, config: AppConfig):
        self.config = config
        self.template_wb: Optional[Workbook] = None
        self.template_sheet: Optional[Worksheet] = None
        self.generator = OutputFileGenerator(config)

    def load_template(self) -> bool:
        """Загрузить шаблон файла."""
        try:
            if not self.config.template_file or not Path(self.config.template_file).exists():
                logger.error("Файл шаблона не найден")
                return False

            self.template_wb = openpyxl.load_workbook(self.config.template_file)
            self.template_sheet = self.template_wb.active
            logger.info(f"Шаблон загружен: {self.config.template_file}")
            return True
        except Exception as e:
            logger.error(f"Ошибка загрузки шаблона: {e}")
            return False

    def collect_competitor_data(self, competitor: CompetitorConfig) -> Dict[str, Any]:
        """
        Собрать данные конкурента в память (без записи в Excel).

        Returns:
            Словарь: {city_name: {field: value}} для городов, где найдены данные
        """
        city_data: Dict[str, Dict[str, Any]] = {}

        if not competitor.file_path or not Path(competitor.file_path).exists():
            logger.warning(f"Файл не найден: {competitor.file_path}")
            return city_data

        try:
            wb = openpyxl.load_workbook(competitor.file_path)
            sheet = wb.worksheets[0]

            for city_name in self.config.cities.keys():
                row_data = self._find_city_row_data(sheet, city_name, competitor)
                if row_data is not None:
                    city_data[city_name] = row_data

            wb.close()
            logger.info(
                f"{competitor.name}: найдено городов {len(city_data)} "
                f"из {len(self.config.cities)}"
            )
        except Exception as e:
            logger.error(f"Ошибка сбора данных {competitor.name}: {e}")

        return city_data

    def _find_city_row_data(
        self,
        source_sheet: Worksheet,
        city_name: str,
        competitor: CompetitorConfig
    ) -> Optional[Dict[str, Any]]:
        """
        Найти строку города и вернуть данные как словарь полей.
        Возвращает None если город не найден.
        """
        city_col = competitor.source_columns.city
        threshold = competitor.fuzzy_match_threshold
        search_names = self.config.get_city_names(city_name)

        for row_idx in range(1, source_sheet.max_row + 1):
            cell_value = source_sheet[f"{city_col}{row_idx}"].value
            if not cell_value:
                continue

            cell_str = str(cell_value).lower()
            matched = False
            for search_name in search_names:
                similarity = fuzz.WRatio(search_name.lower(), cell_str)
                if similarity >= threshold:
                    matched = True
                    logger.debug(
                        f"Город '{city_name}' найден как '{cell_value}' "
                        f"(вариант: '{search_name}', совпадение: {similarity}%)"
                    )
                    break

            if not matched:
                continue

            if not self._check_special_conditions(source_sheet, row_idx, city_name, competitor):
                continue

            # Собрать значения полей
            offsets = competitor.row_offsets
            src_cols = competitor.source_columns
            mappings = [
                (src_cols.convert,    offsets.row_2, 'convert'),
                (src_cols.minimum_1,  offsets.row_3, 'minimum_1'),
                (src_cols.minimum_2,  offsets.row_4, 'minimum_2'),
                (src_cols.volume,     offsets.row_5, 'volume'),
                (src_cols.weight_100, offsets.row_6, 'weight_100'),
                (src_cols.weight_3000,offsets.row_7, 'weight_3000'),
            ]
            row_data: Dict[str, Any] = {}
            for src_col, row_offset, field in mappings:
                row_data[field] = source_sheet[f"{src_col}{row_idx + row_offset}"].value
            return row_data

        logger.debug(
            f"Город '{city_name}' не найден для {competitor.name} "
            f"(варианты: {search_names})"
        )
        return None

    def process_competitor(self, competitor: CompetitorConfig) -> Dict[str, Any]:
        """
        Обработать файл конкурента (запись уже собранных данных в Excel).

        Returns:
            Словарь с результатами обработки
        """
        result = {
            'success': False,
            'competitor': competitor.name,
            'processed_cities': 0,
            'errors': []
        }

        try:
            if not competitor.file_path or not Path(competitor.file_path).exists():
                result['errors'].append(f"Файл не найден: {competitor.file_path}")
                return result

            # Загрузить файл конкурента
            wb = openpyxl.load_workbook(competitor.file_path)
            sheet = wb.worksheets[0]

            logger.info(f"Обработка конкурента: {competitor.name}")

            # Обработать каждый город
            for city_name in self.config.cities.keys():
                found = self._find_and_copy_city_data(
                    sheet, city_name, competitor
                )
                if found:
                    result['processed_cities'] += 1

            wb.close()
            result['success'] = True
            logger.info(f"Обработано городов: {result['processed_cities']}")

        except Exception as e:
            error_msg = f"Ошибка обработки: {str(e)}"
            result['errors'].append(error_msg)
            logger.error(error_msg)

        return result

    def _find_and_copy_city_data(
        self,
        source_sheet: Worksheet,
        city_name: str,
        competitor: CompetitorConfig
    ) -> bool:
        """
        Найти город в исходном файле и скопировать данные в шаблон.
        Поиск ведётся по основному названию и всем псевдонимам города.

        Returns:
            True если город найден и данные скопированы
        """
        city_col = competitor.source_columns.city
        threshold = competitor.fuzzy_match_threshold

        # Все варианты написания: основное + псевдонимы
        search_names = self.config.get_city_names(city_name)

        for row_idx in range(1, source_sheet.max_row + 1):
            cell_value = source_sheet[f"{city_col}{row_idx}"].value

            if not cell_value:
                continue

            cell_str = str(cell_value).lower()

            # Проверяем совпадение с любым вариантом написания
            matched = False
            for search_name in search_names:
                similarity = fuzz.WRatio(search_name.lower(), cell_str)
                if similarity >= threshold:
                    matched = True
                    logger.debug(
                        f"Город '{city_name}' найден как '{cell_value}' "
                        f"(вариант: '{search_name}', совпадение: {similarity}%)"
                    )
                    break

            if not matched:
                continue

            if not self._check_special_conditions(
                source_sheet, row_idx, city_name, competitor
            ):
                continue

            self._copy_row_data(source_sheet, row_idx, city_name, competitor)
            return True

        logger.debug(f"Город '{city_name}' не найден для {competitor.name} "
                     f"(варианты: {search_names})")
        return False

    def _check_special_conditions(
        self,
        sheet: Worksheet,
        row_idx: int,
        city_name: str,
        competitor: CompetitorConfig
    ) -> bool:
        """Проверить специальные условия для конкретного конкурента/города."""
        if not competitor.special_conditions:
            return True

        # Пример: для Энергии и Владивостока проверить колонку B
        if competitor.name == "Энергия" and city_name == "Владивосток":
            cell_value = sheet[f"B{row_idx}"].value
            if cell_value != "Авто":
                return False

        return True

    def _copy_row_data(
        self,
        source_sheet: Worksheet,
        source_row: int,
        city_name: str,
        competitor: CompetitorConfig
    ):
        """Скопировать данные из исходной строки в целевую."""
        offsets = competitor.row_offsets
        src_cols = competitor.source_columns

        # Копировать каждое поле
        mappings = [
            (src_cols.convert, offsets.row_2, 'convert'),
            (src_cols.minimum_1, offsets.row_3, 'minimum_1'),
            (src_cols.minimum_2, offsets.row_4, 'minimum_2'),
            (src_cols.volume, offsets.row_5, 'volume'),
            (src_cols.weight_100, offsets.row_6, 'weight_100'),
            (src_cols.weight_3000, offsets.row_7, 'weight_3000'),
        ]

        for src_col, row_offset, field in mappings:
            src_cell = f"{src_col}{source_row + row_offset}"
            value = source_sheet[src_cell].value
            self.generator.write_competitor_data(competitor, city_name, field, value)

    def _get_merged_top_left(self, cell_ref: str) -> Optional[str]:
        """Вернуть координату верхней левой ячейки объединения, если есть."""
        if not self.template_sheet:
            return None

        for merged_range in self.template_sheet.merged_cells.ranges:
            if cell_ref in merged_range:
                min_col, min_row, _max_col, _max_row = merged_range.bounds
                return f"{get_column_letter(min_col)}{min_row}"

        return None

    def save_output(self) -> bool:
        """Сохранить результирующий файл."""
        try:
            if not self.template_wb:
                logger.error("Нет данных для сохранения")
                return False

            output_path = self.config.output_file
            if not output_path:
                logger.error("Не указан путь для сохранения")
                return False

            self.template_wb.save(output_path)
            logger.info(f"Файл сохранен: {output_path}")
            return True
        except Exception as e:
            logger.error(f"Ошибка сохранения файла: {e}")
            return False

    def process_all(self, progress_callback=None) -> List[Dict[str, Any]]:
        """
        Обработать всех конкурентов.

        Args:
            progress_callback: функция для отображения прогресса (competitor_name, is_done)

        Returns:
            Список результатов обработки
        """
        results = []

        enabled_competitors = [
            comp for comp in self.config.competitors.values()
            if comp.enabled
        ]

        # ШАГ 1 — собрать данные всех конкурентов в память
        # collected: {competitor_name: {city_name: {field: value}}}
        collected: Dict[str, Dict[str, Dict[str, Any]]] = {}
        for competitor in enabled_competitors:
            if progress_callback:
                progress_callback(competitor.name, False)
            collected[competitor.name] = self.collect_competitor_data(competitor)
            if progress_callback:
                progress_callback(competitor.name, True)

        # ШАГ 2 — построить карту присутствия {city: [competitor, ...]}
        # конкурент включается в город только если у него есть хотя бы одно значение
        city_competitors: Dict[str, List[CompetitorConfig]] = {}
        for city in self.config.cities.keys():
            city_competitors[city] = []
            for competitor in enabled_competitors:
                city_data = collected.get(competitor.name, {}).get(city)
                if city_data and any(v is not None for v in city_data.values()):
                    city_competitors[city].append(competitor)

        # ШАГ 3 — генерировать структуру Excel с учётом присутствия
        if not self.generator.generate(city_competitors=city_competitors):
            logger.error("Не удалось создать выходной файл")
            return results

        self.template_wb = self.generator.wb
        self.template_sheet = self.generator.ws

        self.generator.add_markups_sheet()

        # ШАГ 4 — записать данные в ячейки
        for competitor in enabled_competitors:
            for city, fields in collected.get(competitor.name, {}).items():
                # Пропускаем город, если конкурент туда не включён
                if competitor not in city_competitors.get(city, []):
                    continue
                for field, value in fields.items():
                    self.generator.write_competitor_data(competitor, city, field, value)

            result = {
                'success': True,
                'competitor': competitor.name,
                'processed_cities': len(collected.get(competitor.name, {})),
                'errors': []
            }
            results.append(result)

        # ШАГ 5 — сохранить файл
        if self.generator.save():
            logger.info("Обработка завершена успешно")

        return results

    def preview_data(self, competitor: CompetitorConfig, max_rows: int = 10) -> List[Dict[str, Any]]:
        """
        Предварительный просмотр данных из файла конкурента.

        Args:
            competitor: конфигурация конкурента
            max_rows: максимальное количество строк для просмотра

        Returns:
            Список словарей с данными строк
        """
        preview_data = []

        try:
            if not competitor.file_path or not Path(competitor.file_path).exists():
                return preview_data

            wb = openpyxl.load_workbook(competitor.file_path, read_only=True)
            sheet = wb.worksheets[0]

            src_cols = competitor.source_columns

            for row_idx in range(1, min(max_rows + 1, sheet.max_row + 1)):
                row_data = {
                    'row': row_idx,
                    'city': sheet[f"{src_cols.city}{row_idx}"].value,
                    'convert': sheet[f"{src_cols.convert}{row_idx}"].value,
                    'minimum_1': sheet[f"{src_cols.minimum_1}{row_idx}"].value,
                    'minimum_2': sheet[f"{src_cols.minimum_2}{row_idx}"].value,
                    'volume': sheet[f"{src_cols.volume}{row_idx}"].value,
                    'weight_100': sheet[f"{src_cols.weight_100}{row_idx}"].value,
                    'weight_3000': sheet[f"{src_cols.weight_3000}{row_idx}"].value,
                }
                preview_data.append(row_data)

            wb.close()

        except Exception as e:
            logger.error(f"Ошибка предпросмотра данных: {e}")

        return preview_data

